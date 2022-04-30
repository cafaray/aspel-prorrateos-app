from crypt import methods
from distutils.log import error
from email.policy import default
import functools
import os
from unittest import result
from urllib import response
from urllib.parse import urlparse
from urllib.request import Request

from flask import Flask, redirect, url_for, render_template, request, session, flash, jsonify
import requests

import logging
import logging.config

from requests.models import Response
import json
from jinja2 import TemplateNotFound
from datetime import datetime, timedelta
# import calendar
from openpyxl import Workbook
from decouple import config

import xml.etree.ElementTree as ET

app = Flask(__name__) # , template_folder='templates', static_folder='static')

app.secret_key = "c29ydGUuYmlvdGVjc2EuY29tL2FkbWluCg=="
app.permanent_session_lifetime= timedelta(hours=2)
app.config['EXPLAIN_TEMPLATE_LOADING'] = True
FILE_LOCATION = config('FILE_LOCATION')
BASE_URL_COI = config('BASE_URL_COI')
BASE_URL_SAE = config('BASE_URL_SAE')
BASE_URL_PONDERADOS = config('BASE_URL_PONDERADOS')
UPLOAD_FOLDER = config('UPLOAD_FOLDER' )
LOGGING_FOLDER = config('LOGGING_FOLDER' )
RECORD_STATUS_FOLDER = config('RECORD_STATUS_FOLDER')

date_log = datetime.strftime(datetime.now(), '%Y%m%d')
logging_filename = f'{date_log}_app_prorrateos.log'
logging_filename = os.path.join(LOGGING_FOLDER, logging_filename)
print(' * Log file at:', logging_filename)
logging.basicConfig(filename=logging_filename, 
                    encoding='utf-8', 
                    level=logging.DEBUG, 
                    format="%(asctime)s %(levelname)s %(threadName)s %(name)s %(message)s")

from subprocess import check_output
coi = check_output(['PORT=5000 npm run dev', '../../hello.js'])
print (p)



print(f'''WORKING WITH NEXT PARAMS: 
    * FILE_LOCATION: {FILE_LOCATION}
    * BASE_URL_COI: {BASE_URL_COI}
    * BASE_URL_SAE: {BASE_URL_SAE}
    * BASE_URL_PONDERADOS: {BASE_URL_PONDERADOS}
    * TEMPORARY_FILES: {UPLOAD_FOLDER}
    ''')

status_process=[]

@app.route('/api/gastos/cfdi', methods = ['GET', 'POST'])
def anexa_cfdi():
    try:
        if request.method == 'POST':
            f = request.files['file']
            filename = os.path.join(UPLOAD_FOLDER, f.filename)
            print('file will be saved in:', filename)
            f.save(filename)

            # set search params            
            XML_GASTOS_PREFIJO = config('XML_GASTOS_PREFIJO')
            XML_GASTOS_COMPROBANTE=config('XML_GASTOS_COMPROBANTE')
            XML_GASTOS_EMISOR=config('XML_GASTOS_EMISOR')
            XML_GASTOS_RECEPTOR=config('XML_GASTOS_RECEPTOR')
            XML_GASTOS_COMPLEMENTO=config('XML_GASTOS_COMPLEMENTO')
            XML_GASTOS_COMPLEMENTO_TIMBRE_FISCAL=config('XML_GASTOS_COMPLEMENTO_TIMBRE_FISCAL')
            XML_GASTOS_TIMBRE_FISCAL_PREFIJO=config('XML_GASTOS_TIMBRE_FISCAL_PREFIJO') 
            print(f'''Working with xml search params:
            PREFIX: {XML_GASTOS_PREFIJO}
            COMPROBANTE: {XML_GASTOS_COMPROBANTE}
            EMISOR: {XML_GASTOS_EMISOR}
            RECEPTOR: {XML_GASTOS_RECEPTOR}
            COMPLEMENTO: {XML_GASTOS_COMPLEMENTO}
            TIMBRE_FISCAL_DIGITAL: {XML_GASTOS_COMPLEMENTO_TIMBRE_FISCAL}
            ''')

            # parse an xml file by name
            tree = ET.parse(filename)
            root = tree.getroot()
            # By default, the prefix should be `cfdi`, but it could be other one:
            if 'cfdi:' in root.tag:
                search_prefix_by = 'cfdi:'
            else:
                search_prefix_by = XML_GASTOS_PREFIJO                
            print(f'searching prefix:', search_prefix_by)
            is_cfdi = search_prefix_by in root.tag
            print('isCFDI:',  is_cfdi)
            record_cfdi = {}
            if is_cfdi:
                # document attributes 
                #document_attributes = ['serie', 'folio', 'fecha', 'formaDePago', 'subTotal', 'descuento', 'TipoCambio', 'Moneda', 'total', 'metodoDePago', 'LugarExpedicion']
                #comprobante=root.attrib
                #for attr in document_attributes:
                #    record_cfdi[attr] = comprobante[attr]

                # person attribute
                #person_attributes = ['nombre', 'rfc']
                #emisor = root.find(search_prefix_by+XML_GASTOS_EMISOR).attrib
                #for attr in person_attributes:
                #    record_cfdi['emisor_'+attr] = emisor[attr]
                
                #receptor = root.find(search_prefix_by+XML_GASTOS_RECEPTOR).attrib
                #for attr in person_attributes:
                #    record_cfdi['receptor_'+attr] = receptor[attr]

                # complement attributes: Timbre fiscal digital                
                complement_attribute = ['UUID', 'FechaTimbrado']                
                complemento = root.find(search_prefix_by+XML_GASTOS_COMPLEMENTO)
                tfd = complemento.find(XML_GASTOS_TIMBRE_FISCAL_PREFIJO+XML_GASTOS_COMPLEMENTO_TIMBRE_FISCAL).attrib
                for attr in complement_attribute:
                    record_cfdi[attr] = tfd[attr]
                print('cfdi translated:', record_cfdi)
                return {'status': 'ok', 'cfdi': record_cfdi}
    except Exception as e:
        print('Exception reading file', e)
        return {'status': 'ko', 'message':e.args[0]}


def save_status_process(id, req='', res=''):    
    if (id=='__end'):
        insert_status_process(status_process, req)
    status_process.append({'id': id, 'request': req, 'response': res})

def insert_status_process(status_process, suffix):
    ts = datetime.now().strftime("%Y%m%d%H%M%S")
    filename = f'{ts}_{suffix}.txt'
    print(f'inserting {len(status_process)} records of status process in {filename}')   
    
    file_status = open(os.path.join(RECORD_STATUS_FOLDER, filename), 'a') 
    for s in status_process:
        pref = '-----\n* '
        for k,v in s.items():
            file_status.write(f'{pref}{k}: {v}\n')
            pref = '  '        

@app.route('/api/gastos/polizas', methods=['POST'])
def gastos_polizas():
    status_process = []
    TIPO_POLIZA = 'Dr'
    IMPORTE_CARGO_CERO = 0
    IMPORTE_ABONO_CERO = 0
    CUENTA_IMPUESTO = "1200-001-000"
    DESCRIPCION_IMPUESTO_IVAIEPS = "IVA ACREDITABLE PENDIENTE DE PAGO"

    CUENTA_IMPUESTO_CLAVE = 1
    CUENTA_RETENCION_ISR_CLAVE = 2
    CUENTA_RETENCION_IVA_CLAVE = 3
       
    CUENTA_RETENCIONES_HONORARIOS_ISR = "2150-001-001"
    DESCRIPCION_IMPUESTO_RETENCIONES_HONORARIOS_ISR = "RETENCION 10% ISR"
    
    CUENTA_RETENCIONES_HONORARIOS_IVA = "2150-001-004"
    DESCRIPCION_IMPUESTO_RETENCIONES_HONORARIOS_IVA = "RETENCION 10% IVA"
    
    CUENTA_RETENCIONES_HONORARIOS = "2150-007-000"
    DESCRIPCION_IMPUESTO_RETENCIONES_HONORARIOS = "RETENCION"

    CUENTA_PROVEEDORES = "2110-003-001"
    DESCRIPCION_CUENTA_PROVEEDORES = "PROVEEDORES"

    try:
        payload = request.get_json()
        print('payload received to process invoice: ', payload)
        save_status_process('payload received', payload, 'n/a')
        documento = payload['document']
        if 'cfdi' in payload: 
            uuid = payload['cfdi']['UUID']
            es_cfdi = 1
        else: 
            es_cfdi = 0
            uuid = ''
        detalles = payload['weights']

        fecha_aplicacion = documento['appliedDate']
        df = datetime.strptime(fecha_aplicacion, '%Y-%m-%d')
        ejercicio = df.year
        periodo = df.month
        save_status_process('search folio', f'TIPO_POLIZA={TIPO_POLIZA}, ejercicio={ejercicio}, periodo={periodo}', 'n/a')
        folio = siguiente_folio(TIPO_POLIZA, ejercicio, periodo)
        save_status_process('found folio', f'TIPO_POLIZA={TIPO_POLIZA}, ejercicio={ejercicio}, periodo={periodo}', f'folio: {folio}')
        cargo_total = 0
        if(folio['folio']>0):
            next_folio = folio['folio']
        else:
            next_folio = 'Imposible determinar un folio para: {0} {1}-{2}'.format('Dr', ejercicio, periodo)
            raise Exception('Imposible determinar un folio para: {0} {1}-{2}'.format('Dr', ejercicio, periodo))
            
        dia = datetime.now().day
        descripcionPoliza = f'{documento["description"]} {documento["appliedDate"]} {documento["supplier"]} {documento["reference"]}'
        auxiliares = []
        save_status_process('defining auxiliars', detalles, 'n/a')
        for detalle in detalles:
            cargo = detalle['value']
            cargo_total = cargo_total + cargo
            abono = IMPORTE_ABONO_CERO
            cuenta_contable = detalle['id'][0:12]
            tipo_cambio = documento['exchangeRate']
            unidad_negocio = detalle['id'][15:]
            auxiliares.append({'cuentaContable':cuenta_contable, 'unidadNegocio':unidad_negocio, 'descripcion':descripcionPoliza, 'tipoCambio':tipo_cambio, 'cargo':cargo, 'abono':abono})
        
        # Manejo para los impuestos
        if documento['tax4']>0:
            cargo = documento['tax4']
            cargo_total = cargo_total + cargo
            abono = IMPORTE_ABONO_CERO
            valor_cuenta = impuesto(CUENTA_IMPUESTO_CLAVE)
            if valor_cuenta:
                cuenta_contable = valor_cuenta['cuenta']
                descripcion = valor_cuenta['impuesto']
            else:
                cuenta_contable = CUENTA_IMPUESTO
                descripcion = DESCRIPCION_IMPUESTO_IVAIEPS            
            auxiliares.append({'cuentaContable':cuenta_contable, 'unidadNegocio':'', 'descripcion':descripcion, 'tipoCambio':1, 'cargo':cargo, 'abono':abono})

        if documento['taxes']>0:
            if documento['tax1']>0:
                abono = documento['tax1']
                valor_cuenta = impuesto(CUENTA_RETENCION_ISR_CLAVE)
                if valor_cuenta:
                    cuenta_contable = valor_cuenta['cuenta']
                    descripcion = valor_cuenta['impuesto']
                else:
                    cuenta_contable = CUENTA_RETENCIONES_HONORARIOS_ISR
                    descripcion = DESCRIPCION_IMPUESTO_RETENCIONES_HONORARIOS_ISR
                auxiliares.append({'cuentaContable':cuenta_contable, 'unidadNegocio':'', 'descripcion':descripcion, 'tipoCambio':1, 'cargo':IMPORTE_CARGO_CERO, 'abono':abono})
            if documento['tax2']>0:
                abono = documento['tax2']
                valor_cuenta = impuesto(CUENTA_RETENCION_ISR_CLAVE)
                if valor_cuenta:
                    cuenta_contable = valor_cuenta['cuenta']
                    descripcion = valor_cuenta['impuesto']
                else:
                    cuenta_contable = CUENTA_RETENCIONES_HONORARIOS_ISR
                    descripcion = DESCRIPCION_IMPUESTO_RETENCIONES_HONORARIOS_ISR
                auxiliares.append({'cuentaContable':cuenta_contable, 'unidadNegocio':'', 'descripcion':descripcion, 'tipoCambio':1, 'cargo':IMPORTE_CARGO_CERO, 'abono':abono})

            if documento['tax3']>0:
                abono = documento['tax3']
                valor_cuenta = impuesto(CUENTA_RETENCION_IVA_CLAVE)
                if valor_cuenta:
                    cuenta_contable = valor_cuenta['cuenta']
                    descripcion = valor_cuenta['impuesto']
                else:
                    cuenta_contable = CUENTA_RETENCIONES_HONORARIOS_IVA 
                    descripcion = DESCRIPCION_IMPUESTO_RETENCIONES_HONORARIOS_IVA
                auxiliares.append({'cuentaContable':cuenta_contable, 'unidadNegocio':'', 'descripcion':descripcion, 'tipoCambio':1, 'cargo':IMPORTE_CARGO_CERO, 'abono':abono})
                
        # anexa detalle en cuenta de proveedores 
        # print('====>', cargo_total, '-', documento['taxes'], '  = ', cargo_total - documento['taxes'])
        auxiliares.append({'cuentaContable': CUENTA_PROVEEDORES, 'unidadNegocio':'', 'descripcion':DESCRIPCION_CUENTA_PROVEEDORES, 'tipoCambio':1, 'cargo':IMPORTE_CARGO_CERO, 'abono':cargo_total-documento['taxes']})
        save_status_process('defining auxiliars', f'tax4: {documento["tax4"]}, taxes: {documento["taxes"]}', auxiliares)
        poliza = {
            'numeroFolio': next_folio, 'descripcion':descripcionPoliza, 'tipo':TIPO_POLIZA, 'numeroDia': dia, 'fecha': fecha_aplicacion,
            'periodo': periodo, 'ejercicio': ejercicio, 'cfdi': es_cfdi, 'uuid': uuid,
            'auxiliares': auxiliares
        }
        print('invoice prepared to be inserted:', poliza)
        save_status_process('invoice prepared to be inserted:', poliza, 'n/a')
                
        response = inserta_poliza_coi(poliza)
        save_status_process('invoice prepared to be inserted:', poliza, response)
        print('insert poliza and auxiliar results in:', response) 
        if 'auxiliars' in response and response['auxiliars']>0:
            save_status_process('actualiza_contabilizado:', f'reference: {documento["reference"]}, supplierId: {documento["supplierId"]}, chargeId: {documento["chargeId"]}, conceptId: {documento["conceptId"]}', 'n/a')
            response = actualiza_contabilizado(documento['reference'], documento['supplierId'], documento['chargeId'], documento['conceptId'])
            save_status_process('actualiza_contabilizado:', f'reference: {documento["reference"]}, supplierId: {documento["supplierId"]}, chargeId: {documento["chargeId"]}, conceptId: {documento["conceptId"]}', response)
            save_status_process('actualiza_folio:', f'tipo poliza: {TIPO_POLIZA}, ejercicio: {ejercicio}, periodo: {periodo}, folio: {next_folio}', 'n/a')
            response = actualiza_folio(TIPO_POLIZA, ejercicio, periodo, next_folio)
            save_status_process('actualiza_folio:', f'tipo poliza: {TIPO_POLIZA}, ejercicio: {ejercicio}, periodo: {periodo}, folio: {next_folio}', response)            
            save_status_process('__end', next_folio)
            
            return {'status': 'ok', 'data':payload}
        else:
            return {'status': 'ko', 'message':f'Algo ha ido mal y no se ha insertado la poliza correctamente. {response}'}
    except Exception as e:
        print(f'Exception in {request.method} gastos_polizas:', e)
        save_status_process('__end', next_folio)
        return {'status': 'ko', 'message':e.args[0]}

def inserta_poliza_coi(poliza):
    try:
        url = f'{BASE_URL_COI}polizas'
        poliza_coi = { 
            'TIPO_POLI': poliza['tipo'], 
            'NUM_POLIZ': str(poliza['numeroFolio']), 
            'PERIODO': poliza['periodo'], 
            'EJERCICIO': poliza['ejercicio'], 
            'FECHA_POL': poliza['fecha'], 
            'CONCEP_PO': poliza['descripcion'],
            'NUM_PART': len(poliza['auxiliares']), 
            'TIENEDOCUMENTOS': poliza['cfdi'], 
            'UUID': poliza['uuid'], 
            'PROCCONTAB': poliza['ejercicio'],   
        }
        response = invoke(url, 'POST', json.dumps(poliza_coi), 'application/json')
        save_status_process('inserta_poliza_coi:', '', response)
        if response.ok:
            result =response.json()
            print('response for inserted invoice:', result['NUM_POLIZ'])
            save_status_process('inserta_poliza_coi:', 'poliza_coi', result)
            poliza['numeroFolio'] = result['NUM_POLIZ']
            # add auxiliar
            url = f'{url}/auxiliares'
            i = 1            
            for detalle in poliza['auxiliares']:
                if detalle['abono']>0:
                    debe_haber = 'H'
                else:
                    debe_haber = 'D'
                auxiliar = {
                    'TIPO_POLI': poliza['tipo'],
                    'NUM_POLIZ': poliza['numeroFolio'],
                    'NUM_PART': i,
                    'PERIODO':poliza['periodo'],
                    'EJERCICIO':poliza['ejercicio'],
                    'NUM_CTA':detalle['cuentaContable'],
                    'FECHA_POL': poliza['fecha'],
                    'CONCEP_PO': detalle['descripcion'],
                    'DEBE_HABER': debe_haber,
                    'MONTOMOV': round(detalle['cargo'] + detalle['abono'], 2),
                    'NUMDEPTO': detalle['unidadNegocio'],
                    'TIPOCAMBIO': detalle['tipoCambio'],                
                    'ORDEN':i
                }
                # inserta auxiliar
                save_status_process('inserta_auxiliar_coi:', auxiliar, '')
                aresponse = invoke(url, 'POST', json.dumps(auxiliar), 'application/json')
                save_status_process('inserta_auxiliar_coi:', auxiliar, aresponse)
                if aresponse.ok:
                    print('Auxiliar insertado: ', auxiliar)
                else:
                    print('El auxiliar no fue insertado: ', auxiliar)
                i = i + 1
            return {'status': 'ok', 'id': poliza['numeroFolio'], 'auxiliars': i}
        else:
            raise 'La poliza no ha sido insertada: ' + poliza_coi 
    except Exception as e:
        print(f'Exception in inserta_poliza_coi:', e)
        raise e

def actualiza_folio(tipo_poliza, ejercicio, periodo, folio):
    try:
        payload = {'tipo': tipo_poliza, 'periodo': periodo, 'folio': folio}
        url = f'{BASE_URL_COI}folios/{ejercicio}'
        response = invoke(url, 'PUT', json.dumps(payload), 'application/json')
        if response and response.ok:
            result = response.json()
            print('response:', result)
            return result
        else:
            print('response:', response.status_code)
            return {'status': 'ko'}
    except Exception as e:
        print(f'Exception in actualiza_folio:', e)
        raise e

def actualiza_contabilizado(referencia, clave_proveedor, cargo, concepto):
    try:
        payload = {'reference': referencia, 'supplierId': clave_proveedor, 'chargeId': cargo, 'conceptId': concepto}
        url = f'{BASE_URL_SAE}pagos/'
        response = invoke(url, 'PUT', json.dumps(payload), 'application/json')
        if response and response.ok:
            result = response.json()
            print('response:', result)
            return {'status': 'ok', 'message': result}
        else:
            print('response: ', response.status_code)
            return {'status': 'ko', 'message': response}
    except Exception as e:
        print(f'Exception in actualiz contabilizado:', e)
        raise e

@app.route('/api/impuestos/<id>', methods=['PUT'])
def update_impuesto(id: int):
    try:        
        #payload = {'account':account, 'business_unit_id':businessUnitId, 'weighted': weighted}
        if (request.method=='PUT'):            
            #print('request', id, request.form)
            payload = request.get_json()            
            url = BASE_URL_PONDERADOS+'/cuentas/'+id            
            #print('url:', url)
            #print('payload:', payload)
            response=invoke(url, 'PUT', payload=payload)
            if (response and response.ok):
                result = response.json()
                print('result put impuesto:', result)
                return result
            else:
                print('response', response.status_code)
                return {'status':'ko'}
    except Exception as e:
        print(f'Exception in {request.method} impuesto:', e)
        return e


@app.route('/api/ponderados/<id>/detalles', methods=['POST', 'PUT', 'DELETE', 'GET'])
def update_ponderado(id:int):
    try:        
        #payload = {'account':account, 'business_unit_id':businessUnitId, 'weighted': weighted}
        if (request.method=='POST'):            
            #print('request', id, request.form)
            payload = request.get_json()
            # payload = {'baseAccount': request.form.get('baseAccount'), 'businessAreaId': request.form.get('businessUnit'), 'percentage':request.form.get('percentage') }
            url = BASE_URL_PONDERADOS+'/ponderados/'+id+'/details'
            #print('url:', url)
            #print('payload:', payload)
            response=invoke(url, 'POST', payload=payload)
            if (response and response.ok):
                result = response.json()
                #print('result post porcentajes:', result)
                return result
            else:
                print('response', response.status_code)
                return {'status':'ko'}
        if (request.method=='PUT'):            
            payload = request.get_json()
            url = f'{BASE_URL_PONDERADOS}/ponderados/{id}/details'
            #print('url:', url)
            #print('payload:', payload)
            response=invoke(url, 'PUT', payload=payload)
            if (response and response.ok):
                result = response.json()
                #print('result put porcentajes:', result)
                return result
            else:
                print('response', response.status_code)
                return {'status':'ko'}
        if (request.method=='DELETE'):            
            payload = request.get_json()
            url = f'{BASE_URL_PONDERADOS}/ponderados/{id}/details'
            #print('url:', url)
            #print('payload:', payload)
            response=invoke(url, 'DELETE', payload=payload)
            if (response and response.ok):
                result = response.json()
                #print('result delete porcentaje:', result)
                return result
            else:
                print('response', response.status_code)
                return {'status':'ko'}
        if (request.method=='GET'):             
            # print('headers:\n', request.headers)
            doc = request.headers['X-DOC']
            # print('doc', doc, '\nlen(doc)', len(doc))
            if len(doc)==0:                
                raise Exception('Document not specified')
            #print('getting document:-', doc,'-')
            document = documento_pago(doc)
            #print('found document:', len(document))
            if not document:
                raise Exception('Document not found')
            url = f'{BASE_URL_PONDERADOS}/ponderados/{id}/details'
            response=invoke(url, 'GET')            
            if (response and response.ok):
                result = response.json()  
                if (isinstance(result, list) and result[0]):
                    #print(result)
                    result = aplica_ponderacion(result, document[0])              
                    return jsonify(result)
                else:                    
                    return response.text
            else:
                print('response', response.status_code)
                return {'status':'ko'}
    except Exception as e:        
        print(f'Exception in {request.method} update_ponderado:', e)
        response = Response()
        response.status_code = 500 
        response._content = {"status": "ko", "error": e.args[0]}
        return response

def aplica_ponderacion(ponderados, document):
    try:
        total = 0
        account_prev = ''
        response = list()
        weights = list()
        account_prev = ponderados[0]['account']
        i = 0
        # {'id': 36, 'account': '6500-043-001', 'businessUnitId': 4, 'businessUnit': 'Textil', 'percentage': 1}
        for ponderado in ponderados:
            account = ponderado['account']
            if account!=account_prev:                    
                account_prev = account
                response.append({'account': account_prev, 'weights': weights})
                weights = list()
            business_unit_id = ponderado['businessUnitId']
            business_unit = ponderado['businessUnit']
            percentage = ponderado['percentage']
            ponderacion = document['IMPORTE'] * percentage
            weights.append({'id': f'{account_prev}-X0{str(business_unit_id)}', 'businessUnitId': business_unit_id, 'businessUnit':business_unit, 'percentage': percentage, 'value':ponderacion })
            i = i + 1
            total = total + ponderacion
        response.append({'account': account_prev, 'weights': weights})        #response.append({'account': account_prev, 'wiegths': weights})
        #print(f'account: {account}, value:{ponderacion}, businessUnit:{business_unit}, businessUnitId: {business_unit_id}, percentage:{percentage}')
        #total = round(total, 2)
        result = {'total': total,  'value': document['IMPORTE'], 'document': document['DOCTO'], 'details': response}
        #print('total', total, 'validate:', total==document['IMPORTE'])
        # print(result)
        return result
    except Exception as e:
        print('Some exception while evaluating weights on aplica_ponderacion:', e)
        raise e

def genera_excel_poliza(data):
    try:        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Poliza Dr'
        # Header ->
        ws['A3'] = data['tipo']
        ws['B3'] = data['numero']
        ws['C3'] = data['concepto']
        ws['D3'] = data['dia']
        # Header <-

        # details ->
        j = 4
        for detalle in data['detalles']:
            ws['A'+str(j)] = ''
            ws['B'+str(j)] = detalle['cuenta_contable']
            ws['C'+str(j)] = detalle['departamento']
            ws['D'+str(j)] = detalle['concepto']
            ws['E'+str(j)] = detalle['tipo_cambio']
            if detalle['cargo']>0:
                ws['F'+str(j)] = detalle['cargo']
            else:
                ws['F'+str(j)] = ''
            if detalle['abono']>0:
                ws['G'+str(j)] = detalle['abono']
            else:
                ws['G'+str(j)] = ''
            j = j + 1
        # ->
        ws['B'+str(j)] = 'FIN_PARTIDAS'
        filename = f"{FILE_LOCATION}poliza_diario_venta_{data['numero']}.xlsx"
        wb.save(filename = filename)
        return filename
    except Exception as e:
        print('Error while generar_archivo:', e)
        raise e

def genera_poliza_diario(documents, date_from, number_invoice):
    try:
        TIPO_POLIZA = "Dr"
        DEPARTAMENTO_DEFAULT = 0
        CUENTA_CLIENTE_DEFAULT = "1150-003-000"
        CUENTA_DEPARTAMENTO = "4100-001-000"
        CUENTA_DEPARTAMENTO_IMPUESTO_0 = "4100-001-000"
        CUENTA_IMPUESTO = "2170-001-000"
        CUENTA_IMPUESTO_0 = "2170-002-000"

        concepto = "Ventas del dia " + date_from.strftime("%d-%m-%Y")
        folio = number_invoice
        day = datetime.now().day
        poliza = {'tipo': TIPO_POLIZA, 'numero': folio, 'concepto': concepto, 'dia': day, 'detalles': []}
        detalles = []
        # detalles: por cada documento
        x = 1
        for document in documents:
            detalle_cliente, detalle_departamento, detalle_impuesto = {}, {}, {} # {'cuenta_contable':'', 'departamento':'', 'concepto': '', 'tipo_cambio':0, 'cargo':0, 'abono':0}
            cuenta_contable = document['CUENTA_CONTABLE']
            if cuenta_contable!='':
                detalle_cliente['cuenta_contable'] = cuenta_contable
            else:
                detalle_cliente['cuenta_contable'] = CUENTA_CLIENTE_DEFAULT
            detalle_cliente['departamento'] = 0
            detalle_cliente['concepto'] = document['NOMBRE'] + " - " + document['CVE_CLPV'] + " Doc. " + document['CVE_DOC']
            detalle_cliente['tipo_cambio'] = document['TIPO_CAMBIO']
            detalle_cliente['cargo'] = document['IMPORTE']
            detalle_cliente['abono'] = 0
            detalles.append(detalle_cliente)

            if document['IVA'] > 0:
                detalle_departamento['cuenta_contable'] = CUENTA_DEPARTAMENTO
            else:
                detalle_departamento['cuenta_contable'] = CUENTA_DEPARTAMENTO_IMPUESTO_0                    
            detalle_departamento['departamento'] = document['items'][0]['CENTRO_COSTOS'] # first item detailed in document gives the value
            detalle_departamento['concepto'] = "Doc. " + document['IDENTIFICADOR'] + " " + document['FECHA_DOC'] + " " + document['NOMBRE']
            detalle_departamento['tipo_cambio'] = document['TIPO_CAMBIO']
            detalle_departamento['cargo'] = 0
            detalle_departamento['abono'] = document['TOTAL']
            detalles.append(detalle_departamento)

            if document['IVA'] > 0:
                detalle_impuesto['cuenta_contable'] = CUENTA_IMPUESTO
                detalle_impuesto['concepto'] = "IVA 16% Doc. " + document['CVE_DOC']
            else:
                detalle_impuesto['cuenta_contable'] = CUENTA_IMPUESTO_0
                detalle_impuesto['concepto'] = "IVA 0% Doc. " + document['CVE_DOC']
            detalle_impuesto['departamento'] = 0                
            detalle_impuesto['tipo_cambio'] = document['TIPO_CAMBIO']
            detalle_impuesto['cargo'] = 0
            detalle_impuesto['abono'] = document['IVA']
            detalles.append(detalle_impuesto)
            print(f'\n===> detalles {x}:', detalles)
            x = x + 1

        #itera
        #print('detalles:', detalles)
        poliza['detalles'] = detalles
        return poliza
    except Exception as e:
        print('Error while genera_poliza', e)
        return None

@app.route('/api/facturasdiario', methods=['POST'])
def procesa_diario_ventas():
    try:        
        #payload = {'account':account, 'business_unit_id':businessUnitId, 'weighted': weighted}
        payload = request.get_json()
        #print('payload:', payload)
        date_from = datetime.strptime(payload['dateFrom'], '%Y-%m-%d').date()
        date_to=date_from   
        number_invoice = payload['numberInvoice']
        #print('date to query:', date_from)
        if (request.method=='POST'):
            documents = documentos_venta_sae(date_from, date_to)
            if len(documents)>0:
                poliza = genera_poliza_diario(documents, date_from, number_invoice)
                if (poliza):
                    print("\n\n\nRecord:\n", poliza, "\n\n\n")
                    # ===> genera el archivo excel:
                    filename = genera_excel_poliza(poliza)
                    if filename:                        
                        print('File created' + filename)            
                        
                        ejercicio = date_from.year
                        periodo = date_from.month
                        #print('ejercicio: ', ejercicio, 'periodo:', periodo)
                        print(f'Updating invoice number to {number_invoice} under {ejercicio}-{periodo}')
                        url = f'{BASE_URL_COI}folios/{ejercicio}'
                        payload = json.dumps({"periodo": periodo, "tipo": "Dr", "folio": number_invoice})
                        response = invoke(url, 'PUT', payload=payload) #, content_type="'Content-Type': 'application/json'")
                        print('response invoke update folio', response)
                        if (response and response.status_code and response.ok):                            
                            result = response.json()
                            return {'status': 'ok', 'details': f'Se ha generado el archivo {filename}'}
                        else:
                            return {'status': 'ko', 'details': response.status_code}
                    else:
                        return {'status': 'ko', 'details': 'Error generating the file'}
                else:
                    return {'status': 'ko', 'details': 'Error generating the structured data for the invoice'}
            else:
                return {'status': 'ko', 'details': 'Error generating the structured data for the invoice'}
        else:
            return {'status': 'ko', 'details': 'Error getting documents'}
    except Exception as e:
        print(f'Exception in {request.method} procesa_diario_ventas:', e)
        return {'status': 'ko', 'details': f'Exception in method {str(e)}'}

def genera_poliza_venta(documents, date_from, date_to, number_invoice):
    try:
        TIPO_POLIZA = "Dr"
        CUENTA_IMPORTE_DEBE = "5000-001-001"
        CUENTA_IMPORTE_HABER = "1190-005-000"

        wb = Workbook()
        ws = wb.active
        ws.title = 'Poliza Dr'
        # Header ->
        ws['A3'] = TIPO_POLIZA
        ws['B3'] = number_invoice
        ws['C3'] = f'Costo de ventas del periodo: {date_from.strftime("%d-%m-%Y")} - {date_to.strftime("%d-%m-%Y")}'
        ws['D3'] = datetime.now().day
        # Header <-

        # detalles: por cada documento
        j = 4        
        for document in documents:
            ws['A'+str(j)] = ''
            
            ws['B'+str(j)] = CUENTA_IMPORTE_DEBE
            ws['C'+str(j)] = document['items'][0]['CENTRO_COSTOS']
            ws['D'+str(j)] = f'Doc. {document["IDENTIFICADOR"]}, {document["FECHA_DOC"]}    {document["CVE_CLPV"]} - {document["NOMBRE"]}'
            ws['E'+str(j)] = document['TIPO_CAMBIO']            
            ws['F'+str(j)] = document['IMPORTE']
            ws['G'+str(j)] = 0
            j = j + 1

            for item in document['items']:
                
                ws['B'+str(j)] = CUENTA_IMPORTE_HABER
                ws['C'+str(j)] = item['CENTRO_COSTOS']
                ws['D'+str(j)] = f'Costo de ventas. Doc. {document["CVE_DOC"]} {document["FECHA_DOC"]} {item["CVE_ART"]}-{item["DESCRIPCION"]}'
                ws['E'+str(j)] = item['TIPO_CAMBIO']            
                ws['F'+str(j)] = 0
                ws['G'+str(j)] = item['IMPORTE']
                j = j + 1
        # ->        
        ws['B'+str(j)] = 'FIN_PARTIDAS'        
        filename = f'{FILE_LOCATION}polizas_venta_{number_invoice}.xlsx'
        wb.save(filename = filename)
        return filename
    except Exception as e:
        print('Error while generar_archivo:', e)
        raise e

@app.route('/api/facturas', methods=['POST'])
def procesa_ventas():
    try:        
        #payload = {'account':account, 'business_unit_id':businessUnitId, 'weighted': weighted}
        payload = request.get_json()
        #print('payload:', payload)
        date_from = datetime.strptime(payload['dateFrom'], '%Y-%m-%d').date()
        date_to = datetime.strptime(payload['dateTo'], '%Y-%m-%d').date()
        number_invoice = payload['numberInvoice']
        #print('date to query:', date_from)
        if (request.method=='POST'):
            #print('requiring documents')
            documents = documentos_venta_sae(date_from, date_to)
            for document in documents:
                sum_haber = 0
                for item in document['items']:
                    sum_haber = sum_haber + item['IMPORTE']
                #print('sum of partida.importe0', sum_haber)
                document['IMPORTE'] = round(sum_haber, 2)
            #print('gatered documents:', len(documents))
            if len(documents)>0:
                # ===> genera el archivo excel:
                filename = genera_poliza_venta(documents, date_from, date_to, number_invoice)
                if (filename):
                    print('File created' + filename)
                    ejercicio = date_from.year
                    periodo = date_from.month
                    #print('ejercicio: ', ejercicio, 'periodo:', periodo)
                    print(f'Updating invoice number to {number_invoice} under {ejercicio}-{periodo}')
                    url = f'{BASE_URL_COI}folios/{ejercicio}'
                    payload = json.dumps({"periodo": periodo, "tipo": "Dr", "folio": number_invoice})
                    response = invoke(url, 'PUT', payload=payload) #, content_type="'Content-Type': 'application/json'")
                    print('response invoke update folio', response)
                    if (response and response.status_code and response.ok):                            
                        result = response.json()
                        print('result', result)
                        return {'status': 'ok', 'details': f'Se ha generado el archivo {filename}'}
                    else:
                        return {'status': 'ko', 'details': response.status_code}
                else:
                    return {'status': 'ko', 'details': 'Error generating the file'}
            else:
                return {'status': 'ko', 'details': 'Error getting documents'}
    except Exception as e:
        print(f'Exception in {request.method} procesa_ventas:', e)
        return {'status': 'ko', 'details': f'Exception in method {str(e)}'}

@app.route('/api/ponderados/<id>', methods=['GET'])
def valores_ponderados(weight_id:int):
    try:
        url = f'{BASE_URL_PONDERADOS}/ponderados/{weight_id}/details'
        response = invoke(url,'GET')
        if response and response.ok:
            return response.json()
        else:
            print('response', response)
            return []        
    except Exception as e:
        print(f'Exception getting valores_ponderados {weight_id}', e)
        raise e


@app.route('/api/ponderados', methods=['POST', 'DELETE'])
def update_ponderados():
    try:        
        #payload = {'account':account, 'business_unit_id':businessUnitId, 'weighted': weighted}
        if (request.method=='POST'):            
            #print('request', id, request.form)
            payload = request.get_json()            
            url = BASE_URL_PONDERADOS+'/ponderados'
            #print('url:', url)
            #print('payload:', payload)
            response=invoke(url, 'POST', payload=payload)
            if (response and response.ok):
                result = response.json()
                #print('result post porcentaje:', result)
                return result
            else:
                print('response', response.status_code)
                return {'status':'ko'}
        if (request.method=='DELETE'):            
            payload = request.get_json()
            url = BASE_URL_PONDERADOS+'/ponderados'
            #print('url:', url)
            #print('payload:', payload)
            response=invoke(url, 'DELETE', payload=payload)
            if (response and response.ok):
                result = response.json()
                #print('result delete porcentaje:', result)
                return result
            else:
                print('response', response.status_code)
                return {'status':'ko'}
    except Exception as e:
        print(f'Exception in {request.method} update_ponderados:', e)
        return {'status': 'ko', 'details': f'Exception in method {str(e)}'}

### Función en desuso. Se utilizo como parte del ejemplo.
def get_segment(request): 
    try:
        segment = request.path.split('/')[-1]
        if segment == '':
            segment = 'index.html'
        return segment    
    except Exception as e:
        print('Exception while getting segment:', e)
        return None          

def validate_user(email: str, password: str):
    try:
        url_login = BASE_URL_PONDERADOS + '/sec/login'        
        credentials = {'user': email, 'password': password}
        print('calling to loging: ', url_login, credentials)
        response = requests.post(url_login, data=credentials)
        print(response)
        if(response.ok):            
            data = response.json()
            print('===> token generated', data['token'])
            print('===> login response.headers', response.headers['x-user-id'])
            session['token'] = data['token']
            session['user'] = email
            session['userId'] = response.headers['x-user-id']
            return True
        else: 
            return False
    except Exception as e:
        print(e)
        return False

def invoke(url, method, payload=None, content_type='application/json'):
    try:
        token = session.get('token', '')
        user_id = session.get('userId', '')
        if token == '': 
            print('*****     Session expired     *****  ')            
            raise Exception('Session expired, start a new session.')
        else:
            headers = {'x-auth':token, 'userId': user_id, 'Content-Type': content_type}
            if payload:
                print('invoking at', method, url, headers, payload)
                req = requests.Request(method, url, headers=headers, data=payload)
            else:
                print('invoking at', method, url, headers)
                req = requests.Request(method, url, headers=headers)
            prepped = req.prepare()
            print('req', prepped)
            print('payload', payload)
            s = requests.Session()
            response = s.send(prepped, timeout=200)
            print('response.status', response.status_code)
            print('response.ok', response.ok)
            return response
    except Exception as e:
        print(' ** Exception invoking:\n===>', method, url, '\n===> payload:',payload, '\n==========\nError:', e, '\n==========')
        raise e

def porcentajes(args):
    try:
        #print('args', args)
        query_params = '?'
        if (args['supplier']): query_params = query_params + 'supplier=' + args['supplier'] + '&'
        if (args['concept']): query_params = query_params + 'concept=' + args['concept']        
        url = f'{BASE_URL_PONDERADOS}/ponderados{query_params}'
        response = invoke(url, 'GET')
        #print('response of invoke', response)
        if (response and response.ok):
            result = response.json()
            #print('result get porcentajes:', result)
            return result
        else:
            print('response', response.status_code)
            return []
    except Exception as e:
        print('Exception getting porcentajes:', e)
        return e

def ponderado(args):
    try:
        #print('args', args)
        params = ''
        if (args['weighted']): params = args['weighted']        
        url = f'{BASE_URL_PONDERADOS}/ponderados/{params}/details'
        response = invoke(url, 'GET')
        #print('response of invoke', response)
        result = {'total': 0, 'weight': params, 'data': []}
        if (response and response.ok):
            data = response.json()
            if len(data)>0:
                suma = 0
                for i in data:
                    suma = suma + i['percentage']
                suma = round(suma, 2)
                result = {'total': (suma*100), 'weight': params, 'data': data}
                # print('result', result)                
            return result
        else:
            print('response', response.status_code)
            return response
    except Exception as e:
        print('Exception getting ponderado: ', e)
        return e

def unidades():
    try:
        if request.method=='GET':
            # get all unidades
            url = f'{BASE_URL_PONDERADOS}/unidades'
            response = invoke(url, 'GET')
            if (response.ok):
                result = response.json()
                # print('result', result)
                return result
            else:
                print('response.status_code:', response.status_code)
                return []
        elif request.method=='POST':
            url = '{BASE_URL_PONDERADOS}/unidades'
            response = invoke(url, 'POST')
            if (response and response.ok):
                result = response.json()
                # print('result', result)
                return result
            else:
                print('response.status_code:', response.status_code)
                return []

    except Exception as e:
        print(f'Exception {request.method} unidades:', e)
        return e

def conceptos():
    try:
        if request.method=='GET':
            # get all unidades
            url = f'{BASE_URL_SAE}conceptos'
            response = invoke(url, 'GET')
            if (response.ok):
                result = response.json()
                # print('result', result)
                return result
            else:
                print('response.status_code:', response.status_code)
                return []        
    except Exception as e:
        print(f'Exception {request.method} conceptos:', e)
        return e

def cuentas_base():
    try:
        if request.method=='GET':
            url = BASE_URL_COI+'cuentas'
            response = invoke(url, 'GET')
            if (response.ok):
                result = response.json()
                return result
            else:
                print('response.status_code:', response.status_code)
    except Exception as e:
        print(f'Exception {request.method} cuentasBase:', e)
        return e

def tipos_poliza_coi():
    try:
        if request.method=='GET':
            url = BASE_URL_COI+'polizas/tipos'
            response = invoke(url, 'GET')
            if (response.ok):
                result = response.json()
                return result
            else:
                print('response.status_code:', response.status_code)
    except Exception as e:
        print(f'Exception {request.method} cuentasBase:', e)
        return e

def departamentos_coi():
    try:
        if request.method=='GET':
            url = BASE_URL_COI+'deptos'
            response = invoke(url, 'GET')
            if (response.ok):
                result = response.json()
                return result
            else:
                print('response.status_code:', response.status_code)
    except Exception as e:
        print(f'Exception {request.method} departamentos:', e)
        return e

def ultimo_periodo():
    try:
        if request.method=='GET':
            url = BASE_URL_COI+'periodos/ultimo'
            response = invoke(url, 'GET')
            if (response.ok):
                result = response.json()
                return result
            else:
                print('response:', response.status_code)
    except Exception as e:
        print(f'Exception {request.method} ultimo_periodo:', e)
        return e

def impuesto(id):
    try:
        url = f'{BASE_URL_PONDERADOS}/cuentas/{id}'
        response = invoke(url, 'GET')
        #print('response', response)
        if (response.ok):
            result = response.json()
            print(f'result impuestos by id={id} ===> ', result)
            return result[0]
        else:
            print('response.status_code:', response.status_code)
            return None
    except Exception as e:
        print(f'Exception getting impuestos by id={id}:', e)
        return None

def impuestos():
    try:
        if request.method=='GET':
            # get all unidades            
            url = BASE_URL_PONDERADOS+'/cuentas'
            response = invoke(url, 'GET')
            #print('response', response)
            if (response.ok):
                result = response.json()
                # print('result', result)
                return result
            else:
                print('response.status_code:', response.status_code)
                return []
    except Exception as e:
        print('Exception getting impuestos:', e)
        return e    

def proveedores():
    try:
        if request.method=='GET':
            # get all unidades
            url = BASE_URL_SAE+'proveedores'
            response = invoke(url, 'GET')
            if (response.ok):
                result = response.json()
                # print('result', result)
                return result
            else:
                print('response.status_code:', response.status_code)
                return [{"CLAVE":-1, "NOMBRE":"NO SE LOCALIZARON PROVEEDORES"}]        
    except Exception as e:
        print('Exception getting proveedores:', e)
        return e

def documentos_compra_sae(date_from, date_to):
    try:
        if request.method=='GET':
            # get all unidades
            params = '?from={0}&to={1}'.format(date_from, date_to)
            url = f'{BASE_URL_SAE}compras/details{params}'
            response = invoke(url, 'GET')
            if (response.ok):
                result = response.json()
                #print('result', result)
                return result
            else:
                print('response.status_code:', response.status_code)
                return []        
    except Exception as e:
        print('Exception getting documentos_compra_sae:', e)
        return e

def deprecated_documentos_venta_sae(date_from, date_to):
    try:
        if request.method=='GET':
            # get all unidades
            params = '?from={0}&to={1}'.format(date_from, date_to)
            url = f'{BASE_URL_SAE}facturas{params}'
            response = invoke(url, 'GET')
            if (response.ok):
                result = response.json()
                #print('result', result)
                return result
            else:
                print('response.status_code:', response.status_code)
                return []        
    except Exception as e:
        print('Exception getting documentos_ventas_sae:', e)
        return e

def siguiente_folio(tipo, ejercicio, periodo):
    try:
        url = BASE_URL_COI+"folios/{0}/next?tipo={1}&periodo={2}".format(ejercicio, tipo, periodo)
        
        response = invoke(url, 'GET')
        if (response.ok):
            result = response.json()
            print('get next folio:', result)
            return result
        else:
            print('get next folio:', response.status_code)
            return response.json()
    except Exception as e:
        print('Exception getting next folio:', e)
        raise e

def documentos_venta_sae(date_from, date_to):
    try:
        params = '?from={0}&to={1}'.format(date_from, date_to)
        url = f'{BASE_URL_SAE}facturas{params}'
        response = invoke(url, 'GET')        
        if (response and response.ok):                        
            result = response.json()
            #print('result facturas', result)
            return result
        else:
            print('response: ', response.status_code)
            return []        
    except Exception as e:
        print('Exception getting facturas: ', e)
        raise e

def documentos_pago_sae(date_from, supplier):
    try:
        print(f'date_from: {date_from}, supplier: {supplier}')
        supplier_id = -1
        try:            
            supplier_id=int(supplier)
        except ValueError as e:
            print(f'Probably not supplier given, value {supplier} can not be converted to int.\nValue error:', e)
        params = f'?appliedDate={date_from}'
        if (supplier_id>0):
            url = f'{BASE_URL_SAE}pagos/proveedores/{supplier}{params}'
        else:
            url = f'{BASE_URL_SAE}pagos{params}'
        response = invoke(url, 'GET')
        if (response and response.ok):
            result = response.json()            
            return result
        else:
            print('response:', response.status_code)
            return []
    except Exception as e:
        print('Exception getting pagos: ', e)
        raise e

def close_session(token):
    try:
        url_logout = BASE_URL_PONDERADOS + '/sec/logout'        
        data = {'session': token}
        print('calling to logout: ', url_logout, data)
        response = requests.post(url_logout, data=data)
        print(response)
        if(response.status_code==200 or response.status_code==204):
            return True
        else: 
            return False        
    except Exception as e:
        print(e)
        return False

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method=='POST':
        print('Come to Login-POST', request.form)
        print(len(request.form))
        if(len(request.form) < 2): 
            return redirect('home/page-500.html')
        is_session = validate_user(request.form['email'], request.form['password'])
        print('Result from validate user and create session is: ', is_session)
        if is_session:
            session.permanent = False
            user = request.form['email']
            #log.info('===> user logged {}'.format(user))
            session['user'] = user  
            return redirect(url_for('home'))
        else:
            flash("Usuario o contraseña incorrecto!", "danger")
            return render_template("/home/login.html")
    else:
        if 'user' in session:
            return redirect(url_for('home'))
        return render_template('/home/login.html')

@app.route('/unidades', methods=['POST', 'GET'])
def unidades_negocio():
    data = unidades()
    if (data):
        print('sendig data to form:', data) 
    else:
        data = []
    return render_template('home/unidades.html', data= data)


@app.route('/pfacturasdiario', methods=['GET'])
def polizas_ventas():
    try:
        print('params:',request.args)
        date_from = request.args.get('datefrom')    
        if date_from:
            data = documentos_venta_sae(date_from, date_from)
            print('\n\n\ndata: ok')
            print('\nrequesting next folio ...')
            df = datetime.strptime(date_from, '%Y-%m-%d')
            ejercicio = df.year
            periodo = df.month
            folio = siguiente_folio('Dr', ejercicio, periodo)            
            if(folio['folio']>0):
                next_folio = folio['folio']
            else:
                next_folio = 'Imposible determinarlo para: {0} {1}-{2}'.format('Dr', ejercicio, periodo)
        else:
            data = []
            next_folio = 'valor calculado con la consulta'
            date_from = ''
        return render_template('home/pfacturasdiario.html', data=data, next_folio=next_folio, date_from=date_from)
    except Exception as e:
        print('some error ocurrs:', e)
        return render_template('home/page-500.html', error=e)

@app.route('/pcompras', methods=['GET'])
def polizas_compras():
    # print('params:',request.args)
    date_from = request.args.get('datefrom')
    date_to = request.args.get('dateto')
    if date_from and date_to:
        data = documentos_compra_sae(date_from, date_to)
    else:
        data = []
    doctypescoi = tipos_poliza_coi()
    baseaccounts = cuentas_base()
    doctypessae = [{"TIPO": "C", "DESCRIP":"Compras"}, {"TIPO": "V", "DESCRIP":"Ventas"}]
    divisionscoi = departamentos_coi() 
    lastperiod = ultimo_periodo() 
    return render_template('home/pcompras.html', data=data, doctypescoi=doctypescoi, doctypessae=doctypessae, baseaccounts=baseaccounts, divisionscoi=divisionscoi, lastperiod=lastperiod)

@app.route('/pagos', methods=['GET'])
def ponderacion_pagos():
    try:
        date_from= request.args.get('datefrom') or datetime.now().strftime('%Y-%m-%d')
        supplier = request.args.get('supplier') or -1
        #if not date_from:            
        #    date_from = datetime.now()
        print('date_from, supplier', date_from, supplier)
        data = documentos_pago_sae(date_from, supplier)
        suppliers = proveedores()
        return render_template('home/pagos.html', data=data, suppliers=suppliers, applied_date=date_from, supplier=supplier)
    except Exception as e:
        return render_template('home/page-500.html', error = e)

def documento_pago(document_id):
    try:
        url = f'{BASE_URL_SAE}pagos/{document_id}'
        response = invoke(url,'GET')
        if response and response.ok:
            return response.json()
        else:
            print('response', response)
            return []
    except Exception as e:
        print(f'Exception getting documento_pago {document_id}', e)
        raise e

def valores_ponderacion(supplier_id):
    try:
        url = f'{BASE_URL_PONDERADOS}/ponderados/proveedores/{supplier_id}'
        response = invoke(url,'GET')
        if response and response.ok:
            return response.json()
        else:
            print('response', response)
            return []        
    except Exception as e:
        print(f'Exception getting valores_ponderacion {supplier_id}', e)
        raise e

def valor_ponderacion(weight_id):
    try:
        url = f'{BASE_URL_PONDERADOS}/ponderados/{weight_id}/details'
        response = invoke(url,'GET')
        if response and response.ok:
            return response.json()
        else:
            print('response', response)
            return []        
    except Exception as e:
        print(f'Exception getting valor_ponderacion {weight_id}', e)
        raise e

@app.route('/sincronizar/proveedores', methods=['GET', 'POST'])
def sincronizar_proveedores():
    try:
        if(request.method=='GET'):
            url = f'{BASE_URL_PONDERADOS}/sync/suppliers'
            result = invoke(url, 'GET')
            if(result and result.ok):
                result = result.json()
                print('result')
                return render_template('home/sync_proveedores.html', data=result)
            else:
                print('result:', result)
                return render_template('home/sync_proveedores.html', data=[])
        if(request.method=='POST'):
            payload = request.get_json()
            print('payload', payload)
            url = f'{BASE_URL_PONDERADOS}/sync/suppliers'
            if payload:
                result = invoke(url, 'POST', json.dumps(payload))
            else:
                result = invoke(url, 'POST')
            if(result and result.ok):
                result = result.json()
                print('=== result ok', result)                
            else:
                print('*** result ko', result)
                result = {'status':'ko', 'message': result}
            return result

    except Exception as e:
        print(f'Error at {request.method} sincronizar_proveedores:', e)
        return render_template('home/page-500.html', error=e)

@app.route('/pagos/<document>/ponderado', methods=['GET', 'POST'])
def ponderacion_pago(document):
    try:
        if (request.method=='GET'):
            print(f'document: {document}')
            doc = documento_pago(document)
            print('working with document: ', doc)
            supplier_id = doc[0]['CVE_PROV']
            weights = valores_ponderacion(supplier_id)
            print('weights found ', weights)
            # business_units = unidades()
            # print('business units found ', business_units)
            return render_template('home/pagos_detalle.html', document=doc[0], weigths=weights, weight=[], defined_weight=-1) #, businessUnits=business_units)
        if (request.method=='POST'):
            print('defined_weights:', request.form['defined_weights'])
            print(f'document: {document}')
            doc = documento_pago(document)            
            supplier_id = doc[0]['CVE_PROV']
            print('supplier_id', supplier_id)
            weight_id = request.form['defined_weights']
            print(f'working with weight-supplier:', weight_id, supplier_id)
            weight = valor_ponderacion(weight_id)
            print('weight found ', weight)            
            weight_applied = aplica_ponderacion(weight, doc[0])
            print('ponderacion aplicada', weight_applied)
            weights = valores_ponderacion(supplier_id)
            # print('weights found ', weights)
            return render_template('home/pagos_detalle.html', document=doc[0], weight=weight_applied, defined_weight=int(weight_id), weigths=weights)
    except Exception as e:
        return render_template('home/page-500.html', error=e)

@app.route('/pfacturas', methods=['GET'])
def polizas_facturas():
    try:
        # print('params:',request.args)
        date_from = request.args.get('datefrom')
        date_to = request.args.get('dateto')
        if date_from and date_to:
            data = documentos_venta_sae(date_from, date_to)
            for document in data:
                sum_haber = 0
                for item in document['items']:
                    sum_haber = sum_haber + item['IMPORTE']
                #print('sum of partida.importe0', sum_haber)
                document['IMPORTE'] = round(sum_haber, 2)
            df = datetime.strptime(date_from, '%Y-%m-%d')
            ejercicio = df.year
            periodo = df.month
            folio = siguiente_folio('Dr', ejercicio, periodo)            
            if(folio['folio']>0):
                next_folio = folio['folio']
            else:
                next_folio = 'Imposible determinarlo para: {0} {1}-{2}'.format('Dr', ejercicio, periodo)
        else:
            data = []
            next_folio = ''
        doctypescoi = tipos_poliza_coi()
        baseaccounts = cuentas_base()
        doctypessae = [{"TIPO": "C", "DESCRIP":"Compras"}, {"TIPO": "V", "DESCRIP":"Ventas"}]
        divisionscoi = departamentos_coi() 
        lastperiod = ultimo_periodo() 
        return render_template('home/pfacturas.html', data=data, doctypescoi=doctypescoi, doctypessae=doctypessae, baseaccounts=baseaccounts, divisionscoi=divisionscoi, lastperiod=lastperiod, next_folio=next_folio, date_from=date_from, date_to=date_to)
    except Exception as e:
        return render_template('home/page-500.html', error=e)

@app.route('/ponderado_detalle', methods=['GET'])
def ponderado_detalle():
    try:
        if (request.method=='GET'):
            # print('request.weighted=',request.args.get('weighted'))
            weighted = request.args.get('weighted')
            data = ponderado({'weighted': weighted})                
            if (data):
                # print('===> getting business units')
                units = unidades()
                # print(units)
                #print('sending data to form:\nbase:', len(data), '\nunits: ', len(units))
            else:
                units = []
                data = []
            return render_template('home/ponderado_detalle.html', data=data, businessUnits=units)
    except Exception as e:
        print('Occurs an exception at ponderado_detalle.GET:', e)
        units = []
        data = []
        return render_template('home/page-500.html', error='Error getting data for ponderado_detalle:'+e)

@app.route('/impuestos', methods=['GET'])
def cuentas_impuesto():        
    data = impuestos()
    if (data):
        print('sending data to form data:', len(data))
    else:
        data=[]
    # print('finally the structure is: ', data)
    return render_template('home/impuestos.html', data=data)

@app.route('/ponderados', methods=['GET'])
def ponderaciones():
    try:
        print('request.supplier=',request.args.get('supplier'))
        supplier = request.args.get('supplier')
        concept = request.args.get('concept')
        data = porcentajes({'supplier': supplier, 'concept':concept})
        suppliers = proveedores()
        concepts = conceptos()
        if (data):
            print('sending data to form data:', len(data), 'suppliers:', len(suppliers), 'concepts', len(concepts))        
        else:
            data=[]
            concepts=[]
            suppliers=[]
        # print('finally the structure is: ', data)
        return render_template('home/ponderados.html', data=data, suppliers=suppliers, concepts=concepts)
    except Exception as e:
        error = 'Exception loading ponderaciones: {err}'.format(err=e)
        print(error)        
        return render_template('home/page-500.html', error=error)

@app.route('/', defaults={'path': 'index.html'})
# @app.route('/<path>')
def home(path):
    print('Come to home', path)
    
    if 'user' in session:
        try:
            return render_template('home/index.html')
        except TemplateNotFound:
            return render_template('home/page-404.html'), 404
    else:
        return redirect(url_for('login'))

@app.route('/logout')
def logout():
    message = "Has salido de la sesión!"
    if ('token' in session):
        token = session['token']
    else:
        token = None
    print('closing session: ', token)
    if (token):        
        not_session = close_session(token)
        print('close session:', not_session)
        message = message + ' Cierre existoso'
    
    session.pop('user', None)
    session.pop('token', None)
    flash(message, "info")
    return redirect(url_for('login'))

if __name__=="__main__":
    #log.info('>>>>> Starting server at http://{}/v1/stats/ <<<<<'.format('0.0.0.0:5000'))  host='0.0.0.0',
    app.run(  port=8080, debug=True )  # debug=True, host='0.0.0.0', port=8080
